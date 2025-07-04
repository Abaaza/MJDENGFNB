import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import cohere
import numpy as np
from openpyxl import load_workbook
import os
import threading
from datetime import datetime

# --- CONFIGURABLE CONSTANTS ---
EMBEDDING_MODEL = "embed-v4.0"
EMBEDDING_BATCH_SIZE = 96        # Cohere supports up to 96 texts per embed call
EMBEDDING_DIMENSION = 1536       # Choose from 256, 512, 1024, or 1536

class PricelistMatcherApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Pricelist Matching Application (Cohere Embed v4.0)")
        self.root.geometry("670x500")

        self.pricelist_path = tk.StringVar()
        self.inquiry_path   = tk.StringVar()
        self.api_key_var    = tk.StringVar()
        self.output_folder  = tk.StringVar()
        self.client         = None

        self.build_widgets()

    def build_widgets(self):
        frm = tk.Frame(self.root, padx=10, pady=10)
        frm.pack(fill=tk.BOTH, expand=True)

        # Pricelist file chooser
        tk.Label(frm, text="Pricelist Excel:").grid(row=0, column=0, sticky="e")
        tk.Entry(frm, textvariable=self.pricelist_path, width=45).grid(row=0, column=1, padx=5)
        tk.Button(frm, text="Browse...",
                  command=lambda: self.pricelist_path.set(
                      filedialog.askopenfilename(
                          title="Select Pricelist File",
                          filetypes=[("Excel files", "*.xlsx *.xls")],
                      )
                  )).grid(row=0, column=2, padx=5, pady=2)

        # Inquiry file chooser
        tk.Label(frm, text="Inquiry Excel:").grid(row=1, column=0, sticky="e")
        tk.Entry(frm, textvariable=self.inquiry_path, width=45).grid(row=1, column=1, padx=5)
        tk.Button(frm, text="Browse...",
                  command=lambda: self.inquiry_path.set(
                      filedialog.askopenfilename(
                          title="Select Inquiry File",
                          filetypes=[("Excel files", "*.xlsx *.xls")],
                      )
                  )).grid(row=1, column=2, padx=5, pady=2)

        # Cohere API key entry
        tk.Label(frm, text="Cohere API Key:").grid(row=2, column=0, sticky="e")
        tk.Entry(frm, textvariable=self.api_key_var, show="*", width=45).grid(row=2, column=1, columnspan=2, sticky="w")

        # Output folder chooser
        tk.Label(frm, text="Output Folder:").grid(row=3, column=0, sticky="e")
        tk.Entry(frm, textvariable=self.output_folder, width=45).grid(row=3, column=1, padx=5)
        tk.Button(frm, text="Select...",
                  command=lambda: self.output_folder.set(
                      filedialog.askdirectory(title="Select Output Folder")
                  )).grid(row=3, column=2, padx=5, pady=2)

        # Progress log box
        tk.Label(frm, text="Progress:").grid(row=4, column=0, sticky="nw", pady=(10,0))
        self.log_box = scrolledtext.ScrolledText(frm, width=85, height=14, state="disabled", wrap=tk.WORD)
        self.log_box.grid(row=4, column=1, columnspan=2, pady=(10,0))

        # Process button
        self.process_btn = tk.Button(
            frm, text="Process", command=self.on_process_thread,
            bg="#4CAF50", fg="white", width=14
        )
        self.process_btn.grid(row=5, column=1, pady=12, sticky="w")

    def log(self, msg):
        self.log_box.config(state="normal")
        self.log_box.insert(tk.END, msg + "\n")
        self.log_box.see(tk.END)
        self.log_box.config(state="disabled")
        self.root.update_idletasks()

    def get_auto_output_path(self):
        folder = self.output_folder.get()
        if not folder:
            raise RuntimeError("Please specify an output folder.")
        now = datetime.now()
        filename = now.strftime("Output_%I-%M-%p_%m-%d-%y.xlsx")
        return os.path.join(folder, filename)

    def on_process_thread(self):
        t = threading.Thread(target=self.on_process)
        t.daemon = True
        t.start()

    def on_process(self):
        self.process_btn.config(state=tk.DISABLED)
        try:
            # Validate inputs
            if not self.pricelist_path.get() or not self.inquiry_path.get():
                raise RuntimeError("Please select both Pricelist and Inquiry files.")
            if not self.api_key_var.get().strip():
                raise RuntimeError("Please enter your Cohere API key.")
            if not self.output_folder.get():
                raise RuntimeError("Please specify an output folder.")

            # Initialize Cohere client
            self.client = cohere.ClientV2(api_key=self.api_key_var.get().strip())
            self.log("Initialized Cohere client.")

            # Determine output path
            output_path = self.get_auto_output_path()
            if os.path.exists(output_path):
                if not messagebox.askyesno("Overwrite?",
                                           f"Output file '{output_path}' exists. Overwrite?"):
                    self.log("Process cancelled by user.")
                    return

            # Load data from Excel
            self.log("Starting processing...")
            price_descs, price_rates = load_pricelist_data(self.pricelist_path.get(), self.log)
            wb_inq, items_to_fill, header_rows = load_inquiry_data(self.inquiry_path.get(), self.log)

            # Match and fill rates
            fill_inquiry_rates(
                client=self.client,
                wb_inq=wb_inq,
                items_to_fill=items_to_fill,
                pricelist_descs=price_descs,
                pricelist_rates=price_rates,
                header_rows=header_rows,
                model=EMBEDDING_MODEL,
                logger_fn=self.log
            )

            # Save and close workbook
            self.log(f"Saving output file as: {output_path}")
            wb_inq.save(output_path)
            wb_inq.close()
            self.log("Output file saved.")
            messagebox.showinfo("Success", f"Pricing completed.\nOutput saved to:\n{output_path}")

        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.log(f"Error: {e}")
        finally:
            self.process_btn.config(state=tk.NORMAL)


# --- CORE PROCESSING FUNCTIONS ---

def preprocess_text(s):
    import re
    if not s:
        return ""
    s = s.lower().strip()
    s = re.sub(r"\s+", " ", s)
    # normalize common units/abbreviations
    s = s.replace("mm.", "mm").replace("cm.", "cm")
    s = s.replace("r.c.c.", "rcc").replace("reinforced cement concrete", "rcc")
    return s

def load_pricelist_data(pricelist_path, logger_fn):
    logger_fn("Reading pricelist file...")
    try:
        wb_price = load_workbook(pricelist_path, read_only=True, data_only=True)
    except Exception as e:
        raise RuntimeError(f"Failed to open pricelist file: {e}")

    descriptions = []
    rates = []
    for sheet in wb_price.worksheets:
        logger_fn(f"Processing pricelist sheet '{sheet.title}'...")
        for row in sheet.iter_rows(values_only=True):
            if not row or all(cell is None for cell in row):
                continue
            desc_val = row[1] if len(row) > 1 else row[0]
            rate_val = row[-1]
            if desc_val and rate_val not in (None, "", 0):
                descriptions.append(preprocess_text(str(desc_val)))
                rates.append(rate_val)
    wb_price.close()
    if not descriptions:
        raise RuntimeError("No item descriptions with rates found in pricelist file.")
    logger_fn(f"Loaded {len(descriptions)} pricelist items.")
    return descriptions, rates

def load_inquiry_data(inquiry_path, logger_fn):
    logger_fn("Reading inquiry file...")
    try:
        wb_inq = load_workbook(inquiry_path, read_only=False, data_only=False)
    except Exception as e:
        raise RuntimeError(f"Failed to open inquiry file: {e}")

    items_to_fill = []
    header_rows  = {}
    for sheet in wb_inq.worksheets:
        logger_fn(f"Scanning inquiry sheet '{sheet.title}' for headers...")
        desc_col = rate_col = qty_col = None
        header_row_idx = None
        for row_idx in range(1, 11):
            for col_idx in range(1, sheet.max_column + 1):
                val = sheet.cell(row=row_idx, column=col_idx).value
                if isinstance(val, str):
                    low = val.strip().lower()
                    if low == "description":
                        desc_col = col_idx
                    elif low == "rate":
                        rate_col = col_idx
                    elif low in ("qty", "quantity"):
                        qty_col = col_idx
            if desc_col and rate_col:
                header_row_idx = row_idx
                header_rows[sheet] = (header_row_idx, desc_col, rate_col, qty_col)
                break
        if not desc_col or not rate_col:
            logger_fn(f"Skipping sheet '{sheet.title}' (no valid headers).")
            continue
        logger_fn((f"Found headers in '{sheet.title}' at row {header_row_idx} "
                  f"(Desc col={desc_col}, Rate col={rate_col}, Qty col={qty_col})."))
        for r in range(header_row_idx + 1, sheet.max_row + 1):
            desc_cell = sheet.cell(row=r, column=desc_col)
            rate_cell = sheet.cell(row=r, column=rate_col)
            qty_cell  = sheet.cell(row=r, column=qty_col) if qty_col else None
            if not desc_cell.value or str(desc_cell.value).strip() == "":
                continue
            if qty_col and (qty_cell.value is None or str(qty_cell.value).strip() == ""):
                continue
            if rate_cell.value not in (None, ""):
                continue
            items_to_fill.append((rate_cell, preprocess_text(str(desc_cell.value))))
        cnt = len([i for i in items_to_fill if i[0].parent == sheet])
        logger_fn(f"Found {cnt} items to price in '{sheet.title}'.")
    if not items_to_fill:
        raise RuntimeError("No inquiry items with empty rates found in the inquiry file.")
    return wb_inq, items_to_fill, header_rows

def get_embeddings(client, texts, model, logger_fn, input_type="search_document"):
    embeddings = []
    for i in range(0, len(texts), EMBEDDING_BATCH_SIZE):
        batch = texts[i : i + EMBEDDING_BATCH_SIZE]
        logger_fn(f"Requesting batch {i//EMBEDDING_BATCH_SIZE + 1} from Cohere...")
        try:
            resp = client.embed(
                texts=batch,
                model=model,
                input_type=input_type,
                output_dimension=EMBEDDING_DIMENSION,
                embedding_types=["float"]
            )
            embeddings.extend(resp.embeddings.float)
        except Exception as e:
            raise RuntimeError(f"Cohere embed API call failed: {e}")
        logger_fn("Received embeddings from Cohere.")
    return np.array(embeddings)

def fill_inquiry_rates(client, wb_inq, items_to_fill,
                        pricelist_descs, pricelist_rates,
                        header_rows, model, logger_fn):
    # Add columns for matched description & similarity
    for sheet, (hdr, _, _, _) in header_rows.items():
        base = sheet.max_column
        sheet.cell(row=hdr, column=base+1, value="Matched Description")
        sheet.cell(row=hdr, column=base+2, value="Similarity Score")

    # Embed pricelist (documents)
    logger_fn("Computing embeddings for pricelist descriptions...")
    pricelist_embeds = get_embeddings(
        client, pricelist_descs, model, logger_fn, input_type="search_document"
    )

    # Embed inquiry (queries)
    logger_fn("Computing embeddings for inquiry descriptions...")
    inquiry_texts   = [desc for (_cell, desc) in items_to_fill]
    inquiry_embeds  = get_embeddings(
        client, inquiry_texts, model, logger_fn, input_type="search_query"
    )

    # Normalize for cosine similarity
    pricelist_norms = np.linalg.norm(pricelist_embeds, axis=1, keepdims=True)
    inquiry_norms   = np.linalg.norm(inquiry_embeds, axis=1, keepdims=True)
    pricelist_unit  = pricelist_embeds / pricelist_norms
    inquiry_unit    = inquiry_embeds   / inquiry_norms

    logger_fn("Calculating similarity scores...")
    sim_matrix = inquiry_unit.dot(pricelist_unit.T)

    # Fill in best match and rate
    for idx, (rate_cell, _) in enumerate(items_to_fill):
        sims       = sim_matrix[idx]
        best_idx   = sims.argmax()
        best_score = float(sims[best_idx])
        best_desc  = pricelist_descs[best_idx]
        best_rate  = pricelist_rates[best_idx]

        sheet      = rate_cell.parent
        matched_c  = sheet.max_column - 1
        score_c    = sheet.max_column

        rate_cell.value = best_rate
        sheet.cell(row=rate_cell.row, column=matched_c).value = best_desc
        sheet.cell(row=rate_cell.row, column=score_c).value = round(best_score, 3)

    logger_fn("All items processed. Best matches and rates filled in.")

if __name__ == "__main__":
    root = tk.Tk()
    app  = PricelistMatcherApp(root)
    root.mainloop()