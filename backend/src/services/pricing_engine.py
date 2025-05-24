"""pricing_engine.py
Reads a master price Excel sheet and returns a lookup dictionary.
This module relies on openpyxl; install via `pip install openpyxl`.
"""

from pathlib import Path
from openpyxl import load_workbook


def load_rates(path: str) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rates = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, rate = row[0], row[1]
        if code:
            rates[str(code).strip()] = float(rate or 0)
    return rates


def apply_rates(boq: list, rates: dict) -> list:
    """Populate unit rates and totals for a list of BoQ items"""
    priced = []
    for item in boq:
        rate = rates.get(item.get('code'))
        unit_rate = item.get('unit_rate') if item.get('unit_rate') is not None else rate
        total = None
        if unit_rate is not None and item.get('qty') is not None:
            total = float(unit_rate) * float(item['qty'])
        priced.append({**item, 'unit_rate': unit_rate, 'total': total})
    return priced

if __name__ == "__main__":
    import json, sys
    rates = load_rates(sys.argv[1])
    items = json.load(open(sys.argv[2]))
    priced = apply_rates(items, rates)
    json.dump(priced, sys.stdout, indent=2)
